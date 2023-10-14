PROMPT = """请严格按照格式输出，每行只能有一项，括号内的字是在该网页上搜索的内容：
示例1
输入：关于环境保护的课程汇报
输出：
应用：
POWERPNT.EXE
WINWORD.EXE
网页：
环境保护
环境保护图片
知网（环境保护）
清华大学网络学堂

示例2
输入：妈妈的生日
输出：
应用：
日历
微信
网页：
淘宝（生日礼物、生日蛋糕）
生日礼物推荐
生日祝福语
美团（餐厅预订）

示例3
输入：看电影
输出：
应用：
爱奇艺
网页：
豆瓣（电影）
电影天堂
猫眼电影
百度云（电影）

示例4
输入：国庆去北京
输出：
应用：
12306
携程
网页：
北京旅游攻略
故宫门票预订
国庆北京景点推荐
北京酒店预订
北京美食推荐

示例5
输入：科研
输出：
应用：
EXCEL.EXE
EndNote
Grammarly
Zotero
Visual Studio Code
网页：
notion
arxiv
谷歌学术
知网
Web of Science
ChatGPT
"""

EXCEL_PROMPT = """注意，只能输出代码！！！！！！
示例：

输入：
职业：游泳老师
任务：开学第一课
应用：Excel

输出：
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font
import os

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()

# 创建学生名单工作表
sheet = wb.active
sheet.title = "学生名单"
headers = ["序号", "姓名", "性别", "电话", "邮箱", "特别注意事项"]
for col_num, header in enumerate(headers, 1):
    col_letter = openpyxl.utils.cell.get_column_letter(col_num)
    sheet[f"{col_letter}1"].font = Font(bold=True)
    sheet[f"{col_letter}1"] = header

# 创建出勤记录工作表
sheet = wb.create_sheet(title="出勤记录")
headers = ["日期", "学生姓名", "是否出勤", "备注"]
for col_num, header in enumerate(headers, 1):
    col_letter = openpyxl.utils.cell.get_column_letter(col_num)
    sheet[f"{col_letter}1"].font = Font(bold=True)
    sheet[f"{col_letter}1"] = header

# 创建学习进度表工作表
sheet = wb.create_sheet(title="学习进度表")
headers = ["学生姓名", "掌握的技能", "次数", "反馈或建议"]
for col_num, header in enumerate(headers, 1):
    col_letter = openpyxl.utils.cell.get_column_letter(col_num)
    sheet[f"{col_letter}1"].font = Font(bold=True)
    sheet[f"{col_letter}1"] = header

# 创建课程计划工作表
sheet = wb.create_sheet(title="课程计划")
headers = ["日期", "主题/内容", "预计时长", "所需设备或道具", "备注"]
for col_num, header in enumerate(headers, 1):
    col_letter = openpyxl.utils.cell.get_column_letter(col_num)
    sheet[f"{col_letter}1"].font = Font(bold=True)
    sheet[f"{col_letter}1"] = header

# 保存工作簿
file_path = "开学第一课.xlsx"
wb.save(file_path)

# 自动打开Excel文件
if os.name == 'nt':  # For Windows
    os.startfile(file_path)
else:  # For MacOS and Linux
    opener = "open" if os.sys.platform == "darwin" else "xdg-open"
    subprocess.call([opener, file_path])
"""

APP_OR_WEB = """这是应用还是文案素材？应用回答1，文案素材回答0。
示例：
输入：携程（机票预订、酒店预订）
输出：1

输入：美团
输出：1

输入：小丑性格分析
输出：0

输入：团队活动方案
输出：0

输入：清华大学网络学堂
输出：1
"""

FILE_TIME_START_END = """请直接写出代码！！！！！！！！
示例：
输入：上周
输出：
now = datetime.datetime.now().date()  
start = (now - datetime.timedelta(days=now.weekday() + 7)).strftime('%Y/%m/%d')
end = start + datetime.timedelta(days=6).strftime('%Y/%m/%d')

输入：本周
输出：
now = datetime.datetime.now().date()
start = (now - datetime.timedelta(days=now.weekday())).strftime('%Y/%m/%d')
end = now.strftime('%Y/%m/%d')

输入：前几天
输出：
now = datetime.datetime.now().date()
start = (now - datetime.timedelta(days=5)).strftime('%Y/%m/%d')
end = start + datetime.timedelta(days=4).strftime('%Y/%m/%d')

输入：三天前
输出：
now = datetime.datetime.now().date()
start = now - datetime.timedelta(days=3).strftime('%Y/%m/%d')
end = start

输入：本学期
输出：
now = datetime.datetime.now().date()
start = (now - datetime.timedelta(days=90)).strftime('%Y/%m/%d')
end = now.strftime('%Y/%m/%d')
"""

OTHER_CODE = """
print(start)
print(end)
fileString = ""
files = search_files_in_time_range((start, end))
for idx, file in enumerate(files):
    file = file.replace("\\\", "/")
    if idx == 0:
        fileString += file
    else:
        fileString += "\\n" + file
print(fileString)
with open("fileString.txt", "w", encoding='utf-8') as f:
    f.write(fileString)
"""

IF_TIME = """判断是否涉及时间。如果涉及，则回答时间；如果不涉及，则输出False；如果时间是以年为单位的，也输出False。

输入：上周的物理实验报告
输出：上周

输入：数学课PPT
输出：False

输入：今天上课的讲义
输出：今天

输入：昨天布置的论文作业
输出：昨天

输入：去年试卷
输出：False

输入：历年考试真题
输出：False

输入：本学期课程大纲
输出：本学期
"""

KEY_WORD_SPLIT = """请把一句话转化成几个重点关键词。示例：
输入：流体力学课资料
输出：
流体
力学
Fluid
Mechanics
CFD

输入：数学课PPT
输出：
数学
线性代数
math

输入：上周的会议纪要
输出：
会议
meeting

输入：实验报告
输出：
实验
报告

输入：大化实验教材
输出：
化学
大化
大学
实验
Chemistry
Experiment

输入：历史研究论文
输出：
历史
研究
论文
History
Research

输入：今天的计算机编程作业
输出：
计算机
编程
Computer
HW
Programming

输入：老师发的经济学讲座视频
输出：
经济学
讲座
视频
Economics
Lecture
"""

FILENAME_EXTENSION = """请写出待查询文件的可能后缀。示例：
输入：数学课PPT
输出：*.pptx | *.pdf

输入：上周的会议纪要
输出：*.docx | *.doc | *.pdf

输入：实验报告
输出：*.pdf | *.docx | *.doc | *.xls | *.xlsx | *.DAT

输入：大化实验教材
输出：*.pdf

输入：历史研究论文
输出：*.pdf | *.docx

输入：今天的计算机编程作业
输出：*.java | *.py | *.pdf | *.pptx

输入：老师发的经济学讲座视频
输出：*.mp4 | *.avi
"""