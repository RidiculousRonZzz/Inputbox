# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment
import os

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()

# 创建实验报告工作表
sheet = wb.active
sheet.title = "实验报告"

# 实验报告标题
sheet.merge_cells("A1:E1")
sheet["A1"] = "实验报告"
sheet["A1"].font = Font(size=18, bold=True)
sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

# 学生信息
sheet["A2"] = "学生姓名"
sheet["B2"] = "学号"
sheet["C2"] = "实验课程"
sheet["D2"] = "班级"
sheet["E2"] = "指导老师"

# 实验信息
sheet["A3"] = "实验日期"
sheet["B3"] = "实验地点"
sheet["C3"] = "实验名称"
sheet["D3"] = "实验成绩"
sheet["E3"] = "备注"

# 实验目的
sheet["A4"] = "实验目的"
sheet.merge_cells("A5:E5")
sheet["A5"].alignment = Alignment(wrap_text=True)

# 实验设备
sheet["A6"] = "实验设备"
sheet.merge_cells("A7:E7")
sheet["A7"].alignment = Alignment(wrap_text=True)

# 实验内容
sheet["A8"] = "实验内容"
sheet.merge_cells("A9:E9")
sheet["A9"].alignment = Alignment(wrap_text=True)

# 实验步骤
sheet["A10"] = "实验步骤"
sheet.merge_cells("A11:E11")
sheet["A11"].alignment = Alignment(wrap_text=True)

# 实验数据及结果
sheet["A12"] = "实验数据及结果"
sheet.merge_cells("A13:E13")
sheet["A13"].alignment = Alignment(wrap_text=True)

# 实验总结
sheet["A14"] = "实验总结"
sheet.merge_cells("A15:E15")
sheet["A15"].alignment = Alignment(wrap_text=True)

# 保存工作簿
file_path = "实验报告.xlsx"
wb.save(file_path)

# 自动打开Excel文件
if os.name == 'nt':  # For Windows
    os.startfile(file_path)
else:  # For MacOS and Linux
    opener = "open" if os.sys.platform == "darwin" else "xdg-open"
    subprocess.call([opener, file_path])